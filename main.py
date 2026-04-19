import sys
from pathlib import Path

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# ====== Configuration ======
INPUT_FILE = "transactions.xlsx"
OUTPUT_FILE = "transactions_updated.xlsx"
PRICE_COLUMN = 3              # Column C
CORRECTED_PRICE_COLUMN = 4    # Column D
CHART_POSITION = "F2"
DISCOUNT_FACTOR = 0.9


def load_workbook(file_path: Path):
    """
    Load an Excel workbook from disk.

    Args:
        file_path: Path to the Excel file.

    Returns:
        A loaded workbook object.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file is not an Excel file.
    """
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if file_path.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx files are supported.")

    return xl.load_workbook(file_path)


def add_corrected_price_header(sheet) -> None:
    """
    Add a header for the corrected price column.

    Args:
        sheet: Active worksheet.
    """
    sheet.cell(row=1, column=CORRECTED_PRICE_COLUMN).value = "Corrected Price"


def calculate_corrected_price(original_price):
    """
    Apply the business rule for price correction.

    Args:
        original_price: Original numeric price.

    Returns:
        Corrected price after discount.

    Raises:
        ValueError: If the price is missing or not numeric.
    """
    if original_price is None:
        raise ValueError("Price cell is empty.")

    if not isinstance(original_price, (int, float)):
        raise ValueError(f"Invalid price value: {original_price}")

    return original_price * DISCOUNT_FACTOR


def update_prices(sheet) -> None:
    """
    Read original prices from the worksheet and write corrected prices
    into a new column.

    Args:
        sheet: Active worksheet.
    """
    # Start at row 2 to skip the header row
    for row in range(2, sheet.max_row + 1):
        original_price = sheet.cell(row=row, column=PRICE_COLUMN).value
        corrected_price = calculate_corrected_price(original_price)

        sheet.cell(row=row, column=CORRECTED_PRICE_COLUMN).value = corrected_price


def build_price_chart(sheet) -> BarChart:
    """
    Build a bar chart from the corrected price column.

    Args:
        sheet: Active worksheet.

    Returns:
        A configured BarChart object.
    """
    values = Reference(
        sheet,
        min_col=CORRECTED_PRICE_COLUMN,
        max_col=CORRECTED_PRICE_COLUMN,
        min_row=1,
        max_row=sheet.max_row,
    )

    chart = BarChart()
    chart.title = "Corrected Prices"
    chart.y_axis.title = "Price"
    chart.x_axis.title = "Transactions"

    # titles_from_data=True uses the first row as the series label
    chart.add_data(values, titles_from_data=True)

    return chart


def add_chart_to_sheet(sheet, chart: BarChart) -> None:
    """
    Insert the chart into the worksheet.

    Args:
        sheet: Active worksheet.
        chart: Chart object to insert.
    """
    sheet.add_chart(chart, CHART_POSITION)


def save_workbook(workbook, output_path: Path) -> None:
    """
    Save the processed workbook to disk.

    Args:
        workbook: Workbook object.
        output_path: Destination file path.
    """
    workbook.save(output_path)


def process_workbook(input_path: Path, output_path: Path) -> None:
    """
    Main workflow for processing the Excel workbook.

    Steps:
    1. Load workbook
    2. Select active sheet
    3. Add corrected price header
    4. Update corrected prices
    5. Create and insert chart
    6. Save updated workbook

    Args:
        input_path: Source workbook path.
        output_path: Destination workbook path.
    """
    workbook = load_workbook(input_path)
    sheet = workbook.active

    add_corrected_price_header(sheet)
    update_prices(sheet)

    chart = build_price_chart(sheet)
    add_chart_to_sheet(sheet, chart)

    save_workbook(workbook, output_path)


def main():
    """
    Application entry point.
    """
    input_path = Path(INPUT_FILE)
    output_path = Path(OUTPUT_FILE)

    try:
        process_workbook(input_path, output_path)
        print(f"Workbook processed successfully: {output_path}")
    except Exception as error:
        print(f"Error: {error}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
